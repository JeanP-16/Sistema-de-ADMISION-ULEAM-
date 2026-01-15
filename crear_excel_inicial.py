"""
Script para crear el archivo Excel inicial del Sistema de Admisión ULEAM
Genera todas las hojas con estructura y algunos datos de ejemplo
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def crear_excel_inicial():
    """Crea el archivo Excel con todas las hojas estructuradas"""
    
    wb = Workbook()
    
    # ========== HOJA 1: REGISTROS NACIONALES ==========
    ws1 = wb.active
    ws1.title = "registros_nacionales"
    
    # Encabezados
    headers1 = ["cedula", "primer_nombre", "segundo_nombre", "apellido_paterno", 
                "apellido_materno", "correo", "celular", "calificacion", 
                "cuadro_honor", "estado", "fecha_registro"]
    
    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(1, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Datos de ejemplo
    registros = [
        ["1316202082", "JEAN PIERRE", "", "FLORES", "PILOSO", 
         "jeanpierre@uleam.edu.ec", "0987654321", 9.5, "SI", "COMPLETO", datetime.now().strftime("%Y-%m-%d")],
        ["1350123456", "BRADDY", "ALEXANDER", "LONDRE", "VERA", 
         "braddy.londre@uleam.edu.ec", "0981234567", 9.2, "SI", "COMPLETO", datetime.now().strftime("%Y-%m-%d")],
        ["1317924551", "BISMARK", "GABRIEL", "CEVALLOS", "SANCHEZ", 
         "bismark.cevallos@uleam.edu.ec", "0976543210", 8.8, "NO", "COMPLETO", datetime.now().strftime("%Y-%m-%d")],
    ]
    
    for row_idx, registro in enumerate(registros, 2):
        for col_idx, valor in enumerate(registro, 1):
            ws1.cell(row_idx, col_idx, valor)
    
    # Ajustar anchos de columna
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 18
    ws1.column_dimensions['E'].width = 18
    ws1.column_dimensions['F'].width = 30
    ws1.column_dimensions['G'].width = 12
    ws1.column_dimensions['H'].width = 12
    ws1.column_dimensions['I'].width = 12
    ws1.column_dimensions['J'].width = 12
    ws1.column_dimensions['K'].width = 15
    
    # ========== HOJA 2: INSCRIPCIONES ==========
    ws2 = wb.create_sheet("inscripciones")
    
    headers2 = ["id_inscripcion", "cedula_postulante", "carrera_id", "carrera_nombre", 
                "jornada", "estado", "fecha_inscripcion"]
    
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(1, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Datos de ejemplo
    inscripciones = [
        [1, "1316202082", 101, "Tecnologías de Información", "Matutina", "CONFIRMADA", datetime.now().strftime("%Y-%m-%d")],
        [2, "1350123456", 102, "Medicina", "Matutina", "CONFIRMADA", datetime.now().strftime("%Y-%m-%d")],
        [3, "1317924551", 101, "Tecnologías de Información", "Vespertina", "CONFIRMADA", datetime.now().strftime("%Y-%m-%d")],
    ]
    
    for row_idx, inscripcion in enumerate(inscripciones, 2):
        for col_idx, valor in enumerate(inscripcion, 1):
            ws2.cell(row_idx, col_idx, valor)
    
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 30
    ws2.column_dimensions['E'].width = 15
    ws2.column_dimensions['F'].width = 15
    ws2.column_dimensions['G'].width = 18
    
    # ========== HOJA 3: EVALUACIONES ==========
    ws3 = wb.create_sheet("evaluaciones")
    
    headers3 = ["id_evaluacion", "cedula_postulante", "nota_verbal", "nota_numerica", 
                "nota_abstracta", "puntaje_total", "estado", "fecha_evaluacion"]
    
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(1, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Datos de ejemplo
    evaluaciones = [
        [1, "1316202082", 9.5, 9.2, 9.0, 850, "EVALUADO", datetime.now().strftime("%Y-%m-%d")],
        [2, "1350123456", 9.0, 9.3, 9.1, 830, "EVALUADO", datetime.now().strftime("%Y-%m-%d")],
        [3, "1317924551", 8.5, 8.8, 8.6, 780, "EVALUADO", datetime.now().strftime("%Y-%m-%d")],
    ]
    
    for row_idx, evaluacion in enumerate(evaluaciones, 2):
        for col_idx, valor in enumerate(evaluacion, 1):
            ws3.cell(row_idx, col_idx, valor)
    
    ws3.column_dimensions['A'].width = 15
    ws3.column_dimensions['B'].width = 18
    ws3.column_dimensions['C'].width = 12
    ws3.column_dimensions['D'].width = 15
    ws3.column_dimensions['E'].width = 15
    ws3.column_dimensions['F'].width = 15
    ws3.column_dimensions['G'].width = 12
    ws3.column_dimensions['H'].width = 18
    
    # ========== HOJA 4: ASIGNACIONES ==========
    ws4 = wb.create_sheet("asignaciones")
    
    headers4 = ["id_asignacion", "cedula_postulante", "carrera_id", "sede_id", 
                "laboratorio", "edificio", "fecha_examen", "hora_inicio", "estado"]
    
    for col, header in enumerate(headers4, 1):
        cell = ws4.cell(1, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Datos de ejemplo
    asignaciones = [
        [1, "1316202082", 101, 1, "LAB-101", "Edificio A - Tecnología", "2025-02-15", "08:00", "ASIGNADO"],
        [2, "1350123456", 102, 1, "LAB-201", "Edificio B - Salud", "2025-02-15", "08:00", "ASIGNADO"],
        [3, "1317924551", 101, 1, "LAB-101", "Edificio A - Tecnología", "2025-02-15", "13:00", "ASIGNADO"],
    ]
    
    for row_idx, asignacion in enumerate(asignaciones, 2):
        for col_idx, valor in enumerate(asignacion, 1):
            ws4.cell(row_idx, col_idx, valor)
    
    ws4.column_dimensions['A'].width = 15
    ws4.column_dimensions['B'].width = 18
    ws4.column_dimensions['C'].width = 12
    ws4.column_dimensions['D'].width = 10
    ws4.column_dimensions['E'].width = 15
    ws4.column_dimensions['F'].width = 25
    ws4.column_dimensions['G'].width = 15
    ws4.column_dimensions['H'].width = 12
    ws4.column_dimensions['I'].width = 12
    
    # ========== HOJA 5: PUNTAJES ==========
    ws5 = wb.create_sheet("puntajes")
    
    headers5 = ["id_puntaje", "cedula_postulante", "nota_bachillerato", "puntaje_senescyt", 
                "bonificacion_merito", "puntaje_final", "porcentaje", "estado_aprobacion"]
    
    for col, header in enumerate(headers5, 1):
        cell = ws5.cell(1, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Datos de ejemplo con cálculos
    puntajes = [
        [1, "1316202082", 9.5, 850, 100, 950, 95.0, "APROBADO"],
        [2, "1350123456", 9.2, 830, 100, 930, 93.0, "APROBADO"],
        [3, "1317924551", 8.8, 780, 0, 780, 78.0, "APROBADO"],
    ]
    
    for row_idx, puntaje in enumerate(puntajes, 2):
        for col_idx, valor in enumerate(puntaje, 1):
            ws5.cell(row_idx, col_idx, valor)
    
    ws5.column_dimensions['A'].width = 12
    ws5.column_dimensions['B'].width = 18
    ws5.column_dimensions['C'].width = 18
    ws5.column_dimensions['D'].width = 18
    ws5.column_dimensions['E'].width = 20
    ws5.column_dimensions['F'].width = 15
    ws5.column_dimensions['G'].width = 12
    ws5.column_dimensions['H'].width = 18
    
    # ========== HOJA 6: ADMINISTRADORES ==========
    ws6 = wb.create_sheet("administradores")
    
    headers6 = ["cedula", "nombre_completo", "rol", "email"]
    
    for col, header in enumerate(headers6, 1):
        cell = ws6.cell(1, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Administrador por defecto (usa una cédula diferente)
    administradores = [
        ["0000000001", "ADMINISTRADOR SISTEMA", "ADMIN", "admin@uleam.edu.ec"],
    ]
    
    for row_idx, admin in enumerate(administradores, 2):
        for col_idx, valor in enumerate(admin, 1):
            ws6.cell(row_idx, col_idx, valor)
    
    ws6.column_dimensions['A'].width = 15
    ws6.column_dimensions['B'].width = 30
    ws6.column_dimensions['C'].width = 10
    ws6.column_dimensions['D'].width = 30
    
    # Guardar archivo
    wb.save("datos_admision.xlsx")
    print("✅ Archivo 'datos_admision.xlsx' creado exitosamente")
    print("\n📊 Hojas creadas:")
    print("  1. registros_nacionales (3 registros)")
    print("  2. inscripciones (3 registros)")
    print("  3. evaluaciones (3 registros)")
    print("  4. asignaciones (3 registros)")
    print("  5. puntajes (3 registros)")
    print("  6. administradores (1 registro)")
    print("\n🔐 Credenciales de prueba:")
    print("  📌 Estudiantes: 1316202082, 1350123456, 1317924551")
    print("  📌 Administrador: 0000000001")

if __name__ == "__main__":
    crear_excel_inicial()
