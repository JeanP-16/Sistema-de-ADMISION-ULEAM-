"""
Gestor de Base de Datos Excel - Sistema de Admisión ULEAM
Maneja todas las operaciones CRUD sobre datos_admision.xlsx
"""

from openpyxl import load_workbook
from datetime import datetime
from threading import Lock
from .validators import CedulaValidator, EmailValidator, CalificacionValidator


class ExcelManager:
    """
    Gestor principal de operaciones Excel
    Implementa CRUD completo con thread-safety
    """
    
    def __init__(self, excel_path="datos_admision.xlsx"):
        self.excel_path = excel_path
        self.lock = Lock()  # Thread-safety para operaciones concurrentes
        self.validators = {
            'cedula': CedulaValidator(),
            'email': EmailValidator(),
            'calificacion': CalificacionValidator()
        }
    
    # ========================================
    # MÉTODOS AUXILIARES
    # ========================================
    
    def _obtener_siguiente_id(self, hoja_nombre):
        """
        Obtiene el siguiente ID auto-incremental para una hoja
        
        Args:
            hoja_nombre: Nombre de la hoja Excel
            
        Returns:
            int: Siguiente ID disponible
        """
        with self.lock:
            wb = load_workbook(self.excel_path)
            ws = wb[hoja_nombre]
            
            max_id = 0
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if row[0] and isinstance(row[0], int):
                    max_id = max(max_id, row[0])
            
            wb.close()
            return max_id + 1
    
    def _formatear_nombre(self, texto):
        """Formatea nombres a mayúsculas y sin espacios extra"""
        if not texto:
            return ""
        return " ".join(texto.upper().strip().split())
    
    # ========================================
    # REGISTROS NACIONALES - CONSULTAS
    # ========================================
    
    def obtener_registro_por_cedula(self, cedula):
        """
        Obtiene un registro completo por cédula
        
        Args:
            cedula: Número de cédula (10 dígitos)
            
        Returns:
            dict: Datos del registro o None si no existe
        """
        # Validar cédula
        es_valida, mensaje = self.validators['cedula'].validar(cedula)
        if not es_valida:
            return None
        
        with self.lock:
            try:
                wb = load_workbook(self.excel_path)
                ws = wb["registros_nacionales"]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if str(row[0]) == str(cedula):
                        registro = {
                            'cedula': str(row[0]),
                            'primer_nombre': row[1] or '',
                            'segundo_nombre': row[2] or '',
                            'apellido_paterno': row[3] or '',
                            'apellido_materno': row[4] or '',
                            'correo': row[5] or '',
                            'celular': str(row[6]) if row[6] else '',
                            'calificacion': float(row[7]) if row[7] else 0.0,
                            'cuadro_honor': row[8] or 'NO',
                            'estado': row[9] or 'PENDIENTE',
                            'fecha_registro': row[10]
                        }
                        wb.close()
                        return registro
                
                wb.close()
                return None
            except Exception as e:
                print(f"Error al obtener registro: {e}")
                return None
    
    def existe_registro(self, cedula):
        """
        Verifica si existe un registro con la cédula dada
        
        Args:
            cedula: Número de cédula
            
        Returns:
            bool: True si existe, False si no
        """
        return self.obtener_registro_por_cedula(cedula) is not None
    
    def listar_todos_registros(self):
        """
        Obtiene lista completa de todos los registros
        
        Returns:
            list: Lista de diccionarios con todos los registros
        """
        with self.lock:
            try:
                wb = load_workbook(self.excel_path)
                ws = wb["registros_nacionales"]
                
                registros = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # Si tiene cédula
                        registro = {
                            'cedula': str(row[0]),
                            'primer_nombre': row[1] or '',
                            'segundo_nombre': row[2] or '',
                            'apellido_paterno': row[3] or '',
                            'apellido_materno': row[4] or '',
                            'correo': row[5] or '',
                            'celular': str(row[6]) if row[6] else '',
                            'calificacion': float(row[7]) if row[7] else 0.0,
                            'cuadro_honor': row[8] or 'NO',
                            'estado': row[9] or 'PENDIENTE',
                            'fecha_registro': row[10]
                        }
                        registros.append(registro)
                
                wb.close()
                return registros
            except Exception as e:
                print(f"Error al listar registros: {e}")
                return []
    
    # ========================================
    # REGISTROS NACIONALES - CRUD
    # ========================================
    
    def insertar_registro(self, datos):
        """
        Inserta un nuevo registro nacional
        
        Args:
            datos: Diccionario con los datos del registro
            
        Returns:
            tuple: (bool éxito, str mensaje)
        """
        # Validar cédula
        es_valida, mensaje = self.validators['cedula'].validar(datos.get('cedula'))
        if not es_valida:
            return False, mensaje
        
        # Verificar que no exista
        if self.existe_registro(datos['cedula']):
            return False, "Ya existe un registro con esta cédula"
        
        # Validar email
        es_valido, mensaje = self.validators['email'].validar(datos.get('correo'))
        if not es_valido:
            return False, mensaje
        
        # Validar calificación
        es_valida, mensaje = self.validators['calificacion'].validar(datos.get('calificacion'))
        if not es_valida:
            return False, mensaje
        
        with self.lock:
            try:
                wb = load_workbook(self.excel_path)
                ws = wb["registros_nacionales"]
                
                # Formatear datos
                nueva_fila = [
                    str(datos['cedula']),
                    self._formatear_nombre(datos.get('primer_nombre', '')),
                    self._formatear_nombre(datos.get('segundo_nombre', '')),
                    self._formatear_nombre(datos.get('apellido_paterno', '')),
                    self._formatear_nombre(datos.get('apellido_materno', '')),
                    datos.get('correo', '').lower().strip(),
                    str(datos.get('celular', '')),
                    float(datos.get('calificacion', 0)),
                    datos.get('cuadro_honor', 'NO').upper(),
                    datos.get('estado', 'PENDIENTE').upper(),
                    datetime.now().strftime("%Y-%m-%d")
                ]
                
                ws.append(nueva_fila)
                wb.save(self.excel_path)
                wb.close()
                
                return True, "Registro insertado exitosamente"
            except Exception as e:
                return False, f"Error al insertar: {str(e)}"
    
    def actualizar_registro(self, cedula, datos):
        """
        Actualiza un registro existente
        
        Args:
            cedula: Cédula del registro a actualizar
            datos: Diccionario con los campos a actualizar
            
        Returns:
            tuple: (bool éxito, str mensaje)
        """
        if not self.existe_registro(cedula):
            return False, "Registro no encontrado"
        
        # Validar email si se proporciona
        if 'correo' in datos:
            es_valido, mensaje = self.validators['email'].validar(datos['correo'])
            if not es_valido:
                return False, mensaje
        
        # Validar calificación si se proporciona
        if 'calificacion' in datos:
            es_valida, mensaje = self.validators['calificacion'].validar(datos['calificacion'])
            if not es_valida:
                return False, mensaje
        
        with self.lock:
            try:
                wb = load_workbook(self.excel_path)
                ws = wb["registros_nacionales"]
                
                # Buscar y actualizar
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    if str(row[0].value) == str(cedula):
                        # Actualizar solo campos proporcionados
                        if 'primer_nombre' in datos:
                            ws.cell(row_idx, 2, self._formatear_nombre(datos['primer_nombre']))
                        if 'segundo_nombre' in datos:
                            ws.cell(row_idx, 3, self._formatear_nombre(datos['segundo_nombre']))
                        if 'apellido_paterno' in datos:
                            ws.cell(row_idx, 4, self._formatear_nombre(datos['apellido_paterno']))
                        if 'apellido_materno' in datos:
                            ws.cell(row_idx, 5, self._formatear_nombre(datos['apellido_materno']))
                        if 'correo' in datos:
                            ws.cell(row_idx, 6, datos['correo'].lower().strip())
                        if 'celular' in datos:
                            ws.cell(row_idx, 7, str(datos['celular']))
                        if 'calificacion' in datos:
                            ws.cell(row_idx, 8, float(datos['calificacion']))
                        if 'cuadro_honor' in datos:
                            ws.cell(row_idx, 9, datos['cuadro_honor'].upper())
                        if 'estado' in datos:
                            ws.cell(row_idx, 10, datos['estado'].upper())
                        
                        wb.save(self.excel_path)
                        wb.close()
                        return True, "Registro actualizado exitosamente"
                
                wb.close()
                return False, "Error al actualizar registro"
            except Exception as e:
                return False, f"Error al actualizar: {str(e)}"
    
    def eliminar_registro(self, cedula):
        """
        Elimina un registro y todas sus dependencias
        
        Args:
            cedula: Cédula del registro a eliminar
            
        Returns:
            tuple: (bool éxito, str mensaje)
        """
        if not self.existe_registro(cedula):
            return False, "Registro no encontrado"
        
        with self.lock:
            try:
                wb = load_workbook(self.excel_path)
                
                # Eliminar de registros_nacionales
                ws_reg = wb["registros_nacionales"]
                for row_idx, row in enumerate(ws_reg.iter_rows(min_row=2), start=2):
                    if str(row[0].value) == str(cedula):
                        ws_reg.delete_rows(row_idx, 1)
                        break
                
                # Eliminar inscripciones relacionadas
                self._eliminar_relacionados(wb, "inscripciones", cedula)
                
                # Eliminar evaluaciones relacionadas
                self._eliminar_relacionados(wb, "evaluaciones", cedula)
                
                # Eliminar asignaciones relacionadas
                self._eliminar_relacionados(wb, "asignaciones", cedula)
                
                # Eliminar puntajes relacionados
                self._eliminar_relacionados(wb, "puntajes", cedula)
                
                wb.save(self.excel_path)
                wb.close()
                
                return True, "Registro y dependencias eliminados exitosamente"
            except Exception as e:
                return False, f"Error al eliminar: {str(e)}"
    
    def _eliminar_relacionados(self, workbook, hoja_nombre, cedula):
        """Elimina registros relacionados de una hoja"""
        ws = workbook[hoja_nombre]
        rows_to_delete = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            # La cédula está en la columna 2 (índice 1)
            if str(row[1].value) == str(cedula):
                rows_to_delete.append(row_idx)
        
        # Eliminar en orden inverso para no afectar índices
        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx, 1)
    
    # ========================================
    # INSCRIPCIONES
    # ========================================
    
    def obtener_inscripcion_por_cedula(self, cedula):
        """Obtiene la inscripción de un postulante"""
        with self.lock:
            try:
                wb = load_workbook(self.excel_path)
                ws = wb["inscripciones"]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if str(row[1]) == str(cedula):
                        inscripcion = {
                            'id_inscripcion': row[0],
                            'cedula_postulante': str(row[1]),
                            'carrera_id': row[2],
                            'carrera_nombre': row[3],
                            'jornada': row[4],
                            'estado': row[5],
                            'fecha_inscripcion': row[6]
                        }
                        wb.close()
                        return inscripcion
                
                wb.close()
                return None
            except Exception as e:
                print(f"Error al obtener inscripción: {e}")
                return None
    
    def insertar_inscripcion(self, datos):
        """Inserta una nueva inscripción"""
        if not self.existe_registro(datos['cedula_postulante']):
            return False, "No existe registro nacional para esta cédula"
        
        with self.lock:
            try:
                wb = load_workbook(self.excel_path)
                ws = wb["inscripciones"]
                
                nuevo_id = self._obtener_siguiente_id("inscripciones")
                
                nueva_fila = [
                    nuevo_id,
                    str(datos['cedula_postulante']),
                    int(datos['carrera_id']),
                    datos['carrera_nombre'],
                    datos['jornada'],
                    datos.get('estado', 'PENDIENTE'),
                    datetime.now().strftime("%Y-%m-%d")
                ]
                
                ws.append(nueva_fila)
                wb.save(self.excel_path)
                wb.close()
                
                return True, "Inscripción creada exitosamente"
            except Exception as e:
                return False, f"Error al insertar inscripción: {str(e)}"
    
    # ========================================
    # EVALUACIONES
    # ========================================
    
    def obtener_evaluacion_por_cedula(self, cedula):
        """Obtiene la evaluación de un postulante"""
        with self.lock:
            try:
                wb = load_workbook(self.excel_path)
                ws = wb["evaluaciones"]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if str(row[1]) == str(cedula):
                        evaluacion = {
                            'id_evaluacion': row[0],
                            'cedula_postulante': str(row[1]),
                            'nota_verbal': float(row[2]) if row[2] else 0.0,
                            'nota_numerica': float(row[3]) if row[3] else 0.0,
                            'nota_abstracta': float(row[4]) if row[4] else 0.0,
                            'puntaje_total': float(row[5]) if row[5] else 0.0,
                            'estado': row[6],
                            'fecha_evaluacion': row[7]
                        }
                        wb.close()
                        return evaluacion
                
                wb.close()
                return None
            except Exception as e:
                print(f"Error al obtener evaluación: {e}")
                return None
    
    def insertar_evaluacion(self, datos):
        """Inserta una nueva evaluación"""
        if not self.existe_registro(datos['cedula_postulante']):
            return False, "No existe registro nacional para esta cédula"
        
        with self.lock:
            try:
                wb = load_workbook(self.excel_path)
                ws = wb["evaluaciones"]
                
                nuevo_id = self._obtener_siguiente_id("evaluaciones")
                
                nueva_fila = [
                    nuevo_id,
                    str(datos['cedula_postulante']),
                    float(datos.get('nota_verbal', 0)),
                    float(datos.get('nota_numerica', 0)),
                    float(datos.get('nota_abstracta', 0)),
                    float(datos.get('puntaje_total', 0)),
                    datos.get('estado', 'PENDIENTE'),
                    datetime.now().strftime("%Y-%m-%d")
                ]
                
                ws.append(nueva_fila)
                wb.save(self.excel_path)
                wb.close()
                
                return True, "Evaluación creada exitosamente"
            except Exception as e:
                return False, f"Error al insertar evaluación: {str(e)}"
    
    # ========================================
    # ASIGNACIONES
    # ========================================
    
    def obtener_asignacion_por_cedula(self, cedula):
        """Obtiene la asignación de un postulante"""
        with self.lock:
            try:
                wb = load_workbook(self.excel_path)
                ws = wb["asignaciones"]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if str(row[1]) == str(cedula):
                        asignacion = {
                            'id_asignacion': row[0],
                            'cedula_postulante': str(row[1]),
                            'carrera_id': row[2],
                            'sede_id': row[3],
                            'laboratorio': row[4],
                            'edificio': row[5],
                            'fecha_examen': row[6],
                            'hora_inicio': str(row[7]) if row[7] else '',
                            'estado': row[8]
                        }
                        wb.close()
                        return asignacion
                
                wb.close()
                return None
            except Exception as e:
                print(f"Error al obtener asignación: {e}")
                return None
    
    def insertar_asignacion(self, datos):
        """Inserta una nueva asignación"""
        if not self.existe_registro(datos['cedula_postulante']):
            return False, "No existe registro nacional para esta cédula"
        
        with self.lock:
            try:
                wb = load_workbook(self.excel_path)
                ws = wb["asignaciones"]
                
                nuevo_id = self._obtener_siguiente_id("asignaciones")
                
                nueva_fila = [
                    nuevo_id,
                    str(datos['cedula_postulante']),
                    int(datos['carrera_id']),
                    int(datos['sede_id']),
                    datos['laboratorio'],
                    datos['edificio'],
                    datos['fecha_examen'],
                    datos.get('hora_inicio', '08:00'),
                    datos.get('estado', 'ASIGNADO')
                ]
                
                ws.append(nueva_fila)
                wb.save(self.excel_path)
                wb.close()
                
                return True, "Asignación creada exitosamente"
            except Exception as e:
                return False, f"Error al insertar asignación: {str(e)}"
    
    # ========================================
    # PUNTAJES
    # ========================================
    
    def obtener_puntaje_por_cedula(self, cedula):
        """Obtiene el puntaje de un postulante"""
        with self.lock:
            try:
                wb = load_workbook(self.excel_path)
                ws = wb["puntajes"]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if str(row[1]) == str(cedula):
                        puntaje = {
                            'id_puntaje': row[0],
                            'cedula_postulante': str(row[1]),
                            'nota_bachillerato': float(row[2]) if row[2] else 0.0,
                            'puntaje_senescyt': int(row[3]) if row[3] else 0,
                            'bonificacion_merito': int(row[4]) if row[4] else 0,
                            'puntaje_final': float(row[5]) if row[5] else 0.0,
                            'porcentaje': float(row[6]) if row[6] else 0.0,
                            'estado_aprobacion': row[7]
                        }
                        wb.close()
                        return puntaje
                
                wb.close()
                return None
            except Exception as e:
                print(f"Error al obtener puntaje: {e}")
                return None
    
    def insertar_puntaje(self, datos):
        """Inserta un nuevo puntaje calculado"""
        if not self.existe_registro(datos['cedula_postulante']):
            return False, "No existe registro nacional para esta cédula"
        
        with self.lock:
            try:
                wb = load_workbook(self.excel_path)
                ws = wb["puntajes"]
                
                nuevo_id = self._obtener_siguiente_id("puntajes")
                
                nueva_fila = [
                    nuevo_id,
                    str(datos['cedula_postulante']),
                    float(datos.get('nota_bachillerato', 0)),
                    int(datos.get('puntaje_senescyt', 0)),
                    int(datos.get('bonificacion_merito', 0)),
                    float(datos.get('puntaje_final', 0)),
                    float(datos.get('porcentaje', 0)),
                    datos.get('estado_aprobacion', 'PENDIENTE')
                ]
                
                ws.append(nueva_fila)
                wb.save(self.excel_path)
                wb.close()
                
                return True, "Puntaje guardado exitosamente"
            except Exception as e:
                return False, f"Error al insertar puntaje: {str(e)}"
    
    # ========================================
    # ADMINISTRADORES
    # ========================================
    
    def es_administrador(self, cedula):
        """Verifica si una cédula corresponde a un administrador"""
        with self.lock:
            try:
                wb = load_workbook(self.excel_path)
                ws = wb["administradores"]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if str(row[0]) == str(cedula) and row[2] == "ADMIN":
                        wb.close()
                        return True
                
                wb.close()
                return False
            except Exception as e:
                print(f"Error al verificar admin: {e}")
                return False
    
    def obtener_info_administrador(self, cedula):
        """Obtiene información de un administrador"""
        with self.lock:
            try:
                wb = load_workbook(self.excel_path)
                ws = wb["administradores"]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if str(row[0]) == str(cedula):
                        admin = {
                            'cedula': str(row[0]),
                            'nombre_completo': row[1],
                            'rol': row[2],
                            'email': row[3]
                        }
                        wb.close()
                        return admin
                
                wb.close()
                return None
            except Exception as e:
                print(f"Error al obtener info admin: {e}")
                return None

    """
    MÉTODOS ADICIONALES para excel_manager.py
    Agrega estos métodos a la clase ExcelManager
    """

    def obtener_todas_sedes(self):
        """
        Obtiene todas las sedes desde la hoja 'sedes'
        
        Returns:
            list: Lista de diccionarios con datos de sedes
        """
        with self.lock:
            try:
                wb = load_workbook(self.excel_path)
                ws = wb["sedes"]
                
                sedes = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # Si tiene ID
                        sede = {
                            'id_sede': row[0],
                            'nombre_sede': row[1] or '',
                            'ciudad': row[2] or '',
                            'direccion': row[3] or '',
                            'telefono': row[4] or '',
                            'capacidad': row[5] or 0,
                            'estado': row[6] or 'ACTIVA'
                        }
                        sedes.append(sede)
                
                wb.close()
                return sedes
            except Exception as e:
                print(f"Error al obtener sedes: {e}")
                return []

    def obtener_sede_por_id(self, id_sede):
        """
        Obtiene una sede específica por ID
        
        Args:
            id_sede: ID de la sede
            
        Returns:
            dict: Datos de la sede o None
        """
        with self.lock:
            try:
                wb = load_workbook(self.excel_path)
                ws = wb["sedes"]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] == id_sede:
                        sede = {
                            'id_sede': row[0],
                            'nombre_sede': row[1] or '',
                            'ciudad': row[2] or '',
                            'direccion': row[3] or '',
                            'telefono': row[4] or '',
                            'capacidad': row[5] or 0,
                            'estado': row[6] or 'ACTIVA'
                        }
                        wb.close()
                        return sede
                
                wb.close()
                return None
            except Exception as e:
                print(f"Error al obtener sede: {e}")
                return None

    def obtener_todas_carreras(self):
        """
        Obtiene todas las carreras desde la hoja 'carreras'
        
        Returns:
            list: Lista de diccionarios con datos de carreras
        """
        with self.lock:
            try:
                wb = load_workbook(self.excel_path)
                ws = wb["carreras"]
                
                carreras = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # Si tiene ID
                        carrera = {
                            'id_carrera': row[0],
                            'nombre_carrera': row[1] or '',
                            'facultad': row[2] or '',
                            'duracion_semestres': row[3] or 0,
                            'cupos_disponibles': row[4] or 0,
                            'modalidad': row[5] or 'PRESENCIAL',
                            'jornadas_disponibles': row[6] or '',
                            'estado': row[7] or 'ACTIVA'
                        }
                        carreras.append(carrera)
                
                wb.close()
                return carreras
            except Exception as e:
                print(f"Error al obtener carreras: {e}")
                return []

    def obtener_carrera_por_id(self, id_carrera):
        """
        Obtiene una carrera específica por ID
        
        Args:
            id_carrera: ID de la carrera
            
        Returns:
            dict: Datos de la carrera o None
        """
        with self.lock:
            try:
                wb = load_workbook(self.excel_path)
                ws = wb["carreras"]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] == id_carrera:
                        carrera = {
                            'id_carrera': row[0],
                            'nombre_carrera': row[1] or '',
                            'facultad': row[2] or '',
                            'duracion_semestres': row[3] or 0,
                            'cupos_disponibles': row[4] or 0,
                            'modalidad': row[5] or 'PRESENCIAL',
                            'jornadas_disponibles': row[6] or '',
                            'estado': row[7] or 'ACTIVA'
                        }
                        wb.close()
                        return carrera
                
                wb.close()
                return None
            except Exception as e:
                print(f"Error al obtener carrera: {e}")
                return None

    def buscar_carreras_por_facultad(self, facultad):
        """
        Busca carreras por nombre de facultad
        
        Args:
            facultad: Nombre de la facultad (puede ser parcial)
            
        Returns:
            list: Lista de carreras que coinciden
        """
        todas_carreras = self.obtener_todas_carreras()
        
        if not facultad:
            return todas_carreras
        
        resultados = []
        facultad_lower = facultad.lower()
        
        for carrera in todas_carreras:
            if facultad_lower in carrera['facultad'].lower():
                resultados.append(carrera)
        
        return resultados

    def obtener_carreras_activas(self):
        """
        Obtiene solo las carreras activas
        
        Returns:
            list: Lista de carreras con estado ACTIVA
        """
        todas_carreras = self.obtener_todas_carreras()
        return [c for c in todas_carreras if c['estado'] == 'ACTIVA']

    def obtener_sedes_activas(self):
        """
        Obtiene solo las sedes activas
        
        Returns:
            list: Lista de sedes con estado ACTIVA
        """
        todas_sedes = self.obtener_todas_sedes()
        return [s for s in todas_sedes if s['estado'] == 'ACTIVA']