"""
Script para crear datos_admision.xlsx con TODAS las hojas
Incluye: registros, inscripciones, evaluaciones, asignaciones, puntajes, administradores, SEDES y CARRERAS
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime


def crear_excel_completo():
    """Crea el archivo Excel con todas las hojas del sistema"""
    
    wb = Workbook()
    
    # Eliminar hoja por defecto
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # ========================================
    # HOJA 1: SEDES
    # ========================================
    print("📍 Creando hoja: sedes")
    ws_sedes = wb.create_sheet("sedes")
    
    # Headers
    headers_sedes = ['id_sede', 'nombre_sede', 'ciudad', 'direccion', 'telefono', 'capacidad', 'estado']
    ws_sedes.append(headers_sedes)
    
    # Formatear headers
    for cell in ws_sedes[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Datos de sedes
    sedes_data = [
        [1, "Sede Matriz Manta", "Manta", "Av. Circunvalación - Vía a San Mateo", "05-2623-740", 5000, "ACTIVA"],
        [2, "Extensión Chone", "Chone", "Calle 10 de Agosto y Bolívar", "05-2695-229", 1500, "ACTIVA"],
        [3, "Extensión Bahía de Caráquez", "Bahía de Caráquez", "Malecón Alberto Santos y Aguilera", "05-2690-030", 1200, "ACTIVA"],
        [4, "Extensión El Carmen", "El Carmen", "Vía Pedernales Km 1.5", "05-2666-175", 800, "ACTIVA"],
        [5, "Extensión Pedernales", "Pedernales", "Calle Eloy Alfaro y 10 de Agosto", "05-2681-319", 600, "ACTIVA"]
    ]
    
    for sede in sedes_data:
        ws_sedes.append(sede)
    
    # Ajustar anchos de columna
    ws_sedes.column_dimensions['A'].width = 12
    ws_sedes.column_dimensions['B'].width = 30
    ws_sedes.column_dimensions['C'].width = 20
    ws_sedes.column_dimensions['D'].width = 40
    ws_sedes.column_dimensions['E'].width = 15
    ws_sedes.column_dimensions['F'].width = 12
    ws_sedes.column_dimensions['G'].width = 12
    
    # ========================================
    # HOJA 2: CARRERAS
    # ========================================
    print("📚 Creando hoja: carreras")
    ws_carreras = wb.create_sheet("carreras")
    
    # Headers
    headers_carreras = ['id_carrera', 'nombre_carrera', 'facultad', 'duracion_semestres', 'cupos_disponibles', 
                        'modalidad', 'jornadas_disponibles', 'estado']
    ws_carreras.append(headers_carreras)
    
    # Formatear headers
    for cell in ws_carreras[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Datos de carreras
    carreras_data = [
        [101, "Tecnologías de la Información", "Facultad de Ciencias Informáticas", 10, 50, 
         "PRESENCIAL", "MATUTINA,VESPERTINA,NOCTURNA", "ACTIVA"],
        [102, "Medicina", "Facultad de Ciencias Médicas", 12, 80, 
         "PRESENCIAL", "MATUTINA", "ACTIVA"],
        [103, "Ingeniería Civil", "Facultad de Ciencias de la Ingeniería", 10, 60, 
         "PRESENCIAL", "MATUTINA,VESPERTINA", "ACTIVA"],
        [104, "Administración de Empresas", "Facultad de Ciencias Administrativas", 8, 70, 
         "PRESENCIAL,SEMIPRESENCIAL", "MATUTINA,VESPERTINA,NOCTURNA", "ACTIVA"],
        [105, "Derecho", "Facultad de Ciencias Jurídicas", 10, 55, 
         "PRESENCIAL", "MATUTINA,NOCTURNA", "ACTIVA"],
        [106, "Enfermería", "Facultad de Ciencias Médicas", 8, 45, 
         "PRESENCIAL", "MATUTINA,VESPERTINA", "ACTIVA"],
        [107, "Psicología", "Facultad de Ciencias Sociales", 10, 40, 
         "PRESENCIAL", "MATUTINA,VESPERTINA", "ACTIVA"],
        [108, "Contabilidad y Auditoría", "Facultad de Ciencias Económicas", 8, 50, 
         "PRESENCIAL,SEMIPRESENCIAL", "MATUTINA,VESPERTINA,NOCTURNA", "ACTIVA"],
        [109, "Agronomía", "Facultad de Ciencias Agropecuarias", 10, 35, 
         "PRESENCIAL", "MATUTINA", "ACTIVA"],
        [110, "Arquitectura", "Facultad de Arquitectura", 10, 40, 
         "PRESENCIAL", "MATUTINA,VESPERTINA", "ACTIVA"]
    ]
    
    for carrera in carreras_data:
        ws_carreras.append(carrera)
    
    # Ajustar anchos de columna
    ws_carreras.column_dimensions['A'].width = 12
    ws_carreras.column_dimensions['B'].width = 35
    ws_carreras.column_dimensions['C'].width = 40
    ws_carreras.column_dimensions['D'].width = 18
    ws_carreras.column_dimensions['E'].width = 18
    ws_carreras.column_dimensions['F'].width = 18
    ws_carreras.column_dimensions['G'].width = 30
    ws_carreras.column_dimensions['H'].width = 12
    
    # ========================================
    # HOJA 3: REGISTROS NACIONALES
    # ========================================
    print("👤 Creando hoja: registros_nacionales")
    ws_registros = wb.create_sheet("registros_nacionales")
    
    # Headers
    headers_registros = ['cedula', 'primer_nombre', 'segundo_nombre', 'apellido_paterno', 'apellido_materno',
                         'correo', 'celular', 'calificacion', 'cuadro_honor', 'estado', 'fecha_registro']
    ws_registros.append(headers_registros)
    
    # Formatear headers
    for cell in ws_registros[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Datos de registros
    registros_data = [
        ['1316202082', 'JEAN PIERRE', '', 'FLORES', 'MENDOZA', 'jean.flores@uleam.edu.ec', 
         '0987654321', 9.5, 'SI', 'COMPLETO', datetime.now().strftime("%Y-%m-%d")],
        ['1350123456', 'BRADDY', 'ALEXANDER', 'LONDRE', 'VERA', 'braddy.londre@uleam.edu.ec', 
         '0991234567', 8.8, 'NO', 'COMPLETO', datetime.now().strftime("%Y-%m-%d")],
        ['1317924551', 'BISMARK', 'GABRIEL', 'CEVALLOS', 'LOOR', 'bismark.cevallos@uleam.edu.ec', 
         '0998765432', 9.2, 'SI', 'COMPLETO', datetime.now().strftime("%Y-%m-%d")]
    ]
    
    for registro in registros_data:
        ws_registros.append(registro)
    
    # Ajustar anchos
    for i, width in enumerate([15, 18, 18, 18, 18, 30, 15, 12, 15, 15, 15], start=1):
        ws_registros.column_dimensions[chr(64 + i)].width = width
    
    # ========================================
    # HOJA 4: INSCRIPCIONES
    # ========================================
    print("📝 Creando hoja: inscripciones")
    ws_inscripciones = wb.create_sheet("inscripciones")
    
    headers_inscripciones = ['id_inscripcion', 'cedula_postulante', 'carrera_id', 'carrera_nombre', 
                             'jornada', 'estado', 'fecha_inscripcion']
    ws_inscripciones.append(headers_inscripciones)
    
    for cell in ws_inscripciones[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    inscripciones_data = [
        [1, '1316202082', 101, 'Tecnologías de la Información', 'MATUTINA', 'CONFIRMADA', datetime.now().strftime("%Y-%m-%d")],
        [2, '1350123456', 103, 'Ingeniería Civil', 'VESPERTINA', 'CONFIRMADA', datetime.now().strftime("%Y-%m-%d")],
        [3, '1317924551', 101, 'Tecnologías de la Información', 'MATUTINA', 'CONFIRMADA', datetime.now().strftime("%Y-%m-%d")]
    ]
    
    for inscripcion in inscripciones_data:
        ws_inscripciones.append(inscripcion)
    
    for i, width in enumerate([15, 20, 12, 35, 15, 15, 18], start=1):
        ws_inscripciones.column_dimensions[chr(64 + i)].width = width
    
    # ========================================
    # HOJA 5: EVALUACIONES
    # ========================================
    print("📊 Creando hoja: evaluaciones")
    ws_evaluaciones = wb.create_sheet("evaluaciones")
    
    headers_evaluaciones = ['id_evaluacion', 'cedula_postulante', 'nota_verbal', 'nota_numerica', 
                            'nota_abstracta', 'puntaje_total', 'estado', 'fecha_evaluacion']
    ws_evaluaciones.append(headers_evaluaciones)
    
    for cell in ws_evaluaciones[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    evaluaciones_data = [
        [1, '1316202082', 9.2, 9.5, 8.8, 920, 'EVALUADO', datetime.now().strftime("%Y-%m-%d")],
        [2, '1350123456', 8.5, 8.8, 8.2, 850, 'EVALUADO', datetime.now().strftime("%Y-%m-%d")],
        [3, '1317924551', 9.0, 9.3, 8.9, 910, 'EVALUADO', datetime.now().strftime("%Y-%m-%d")]
    ]
    
    for evaluacion in evaluaciones_data:
        ws_evaluaciones.append(evaluacion)
    
    for i, width in enumerate([15, 20, 15, 15, 15, 15, 15, 18], start=1):
        ws_evaluaciones.column_dimensions[chr(64 + i)].width = width
    
    # ========================================
    # HOJA 6: ASIGNACIONES
    # ========================================
    print("📍 Creando hoja: asignaciones")
    ws_asignaciones = wb.create_sheet("asignaciones")
    
    headers_asignaciones = ['id_asignacion', 'cedula_postulante', 'carrera_id', 'sede_id', 
                            'laboratorio', 'edificio', 'fecha_examen', 'hora_inicio', 'estado']
    ws_asignaciones.append(headers_asignaciones)
    
    for cell in ws_asignaciones[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    asignaciones_data = [
        [1, '1316202082', 101, 1, 'LAB-A-201', 'Edificio A', '2025-02-15', '08:00', 'ASIGNADO'],
        [2, '1350123456', 103, 1, 'LAB-B-105', 'Edificio B', '2025-02-15', '10:00', 'ASIGNADO'],
        [3, '1317924551', 101, 1, 'LAB-A-201', 'Edificio A', '2025-02-15', '08:00', 'ASIGNADO']
    ]
    
    for asignacion in asignaciones_data:
        ws_asignaciones.append(asignacion)
    
    for i, width in enumerate([15, 20, 12, 10, 15, 15, 15, 12, 15], start=1):
        ws_asignaciones.column_dimensions[chr(64 + i)].width = width
    
    # ========================================
    # HOJA 7: PUNTAJES
    # ========================================
    print("📈 Creando hoja: puntajes")
    ws_puntajes = wb.create_sheet("puntajes")
    
    headers_puntajes = ['id_puntaje', 'cedula_postulante', 'nota_bachillerato', 'puntaje_senescyt', 
                        'bonificacion_merito', 'puntaje_final', 'porcentaje', 'estado_aprobacion']
    ws_puntajes.append(headers_puntajes)
    
    for cell in ws_puntajes[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    puntajes_data = [
        [1, '1316202082', 9.5, 920, 50, 970, 97.0, 'APROBADO'],
        [2, '1350123456', 8.8, 850, 0, 850, 85.0, 'APROBADO'],
        [3, '1317924551', 9.2, 910, 50, 960, 96.0, 'APROBADO']
    ]
    
    for puntaje in puntajes_data:
        ws_puntajes.append(puntaje)
    
    for i, width in enumerate([12, 20, 18, 18, 20, 15, 12, 18], start=1):
        ws_puntajes.column_dimensions[chr(64 + i)].width = width
    
    # ========================================
    # HOJA 8: ADMINISTRADORES
    # ========================================
    print("👨‍💼 Creando hoja: administradores")
    ws_admins = wb.create_sheet("administradores")
    
    headers_admins = ['cedula', 'nombre_completo', 'rol', 'email']
    ws_admins.append(headers_admins)
    
    for cell in ws_admins[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    admins_data = [
        ['0000000001', 'SISTEMA ADMINISTRADOR', 'ADMIN', 'admin@uleam.edu.ec']
    ]
    
    for admin in admins_data:
        ws_admins.append(admin)
    
    for i, width in enumerate([15, 30, 12, 30], start=1):
        ws_admins.column_dimensions[chr(64 + i)].width = width
    
    # ========================================
    # GUARDAR ARCHIVO
    # ========================================
    filename = "datos_admision.xlsx"
    wb.save(filename)
    print(f"\n✅ Archivo '{filename}' creado exitosamente con {len(wb.sheetnames)} hojas:")
    for sheet_name in wb.sheetnames:
        print(f"   - {sheet_name}")
    print(f"\n📊 Estadísticas:")
    print(f"   - Sedes: {len(sedes_data)}")
    print(f"   - Carreras: {len(carreras_data)}")
    print(f"   - Registros: {len(registros_data)}")
    print(f"   - Inscripciones: {len(inscripciones_data)}")
    print(f"   - Evaluaciones: {len(evaluaciones_data)}")
    print(f"   - Asignaciones: {len(asignaciones_data)}")
    print(f"   - Puntajes: {len(puntajes_data)}")
    print(f"   - Administradores: {len(admins_data)}")


if __name__ == "__main__":
    print("=" * 60)
    print("📁 CREANDO ARCHIVO EXCEL COMPLETO - SISTEMA ULEAM")
    print("=" * 60)
    print()
    crear_excel_completo()
    print()
    print("=" * 60)
    print("✅ PROCESO COMPLETADO")
    print("=" * 60)
